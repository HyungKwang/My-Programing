#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Read opensm-statistics.dump (CSV from OpenSM) and write an Excel workbook
with a data sheet and a Charts sheet (line charts per numeric column),
using openpyxl.

Changes:
- Use actual TIME_STAMP values on the X-axis
- Remove legend on the right side
- Remove vertical axis title on the left side
- Keep Y-axis tick labels visible
"""

from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

BASE_YEAR = 2000

MONTHS = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "sept": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}

def is_number(value):
    return isinstance(value, (int, float)) and value is not None

def parse_timestamp(raw: str, default_year: int = BASE_YEAR):
    value, _has_year = parse_timestamp_with_year_info(raw, default_year)
    return value

def parse_timestamp_with_year_info(raw: str, default_year: int = BASE_YEAR):
    s = raw.strip()
    if not s:
        return None, False

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y/%m/%d %H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(s, fmt), True
        except ValueError:
            pass

    parts = s.split()
    if len(parts) == 3:
        month_text, day_text, time_text = parts
        year = default_year
        has_year = False
    elif len(parts) == 4 and parts[2].isdigit():
        month_text, day_text, year_text, time_text = parts
        year = int(year_text)
        has_year = True
    else:
        return s, False

    month = MONTHS.get(month_text.lower())
    if month is None:
        return s, False

    try:
        hour_text, minute_text, second_text = time_text.split(":")
        timestamp = datetime(
            year,
            month,
            int(day_text),
            int(hour_text),
            int(minute_text),
            int(second_text),
        )
    except ValueError:
        return s, False

    return timestamp, has_year

def add_years(value: datetime, years: int):
    try:
        return value.replace(year=value.year + years)
    except ValueError:
        return value.replace(year=value.year + years, day=28)

def format_timestamp_for_excel(value):
    if isinstance(value, datetime):
        return value.strftime("%b %d %H:%M:%S")
    return value

def parse_cell(raw: str, is_timestamp: bool = False):
    s = raw.strip()
    if not s:
        return None

    if is_timestamp:
        return parse_timestamp(s)

    try:
        return int(s)
    except ValueError:
        pass

    try:
        return float(s)
    except ValueError:
        return s

def read_dump_csv(path: Path):
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, skipinitialspace=True)
        rows_iter = iter(reader)

        try:
            header = next(rows_iter)
        except StopIteration:
            return [], []

        headers = [h.strip() for h in header]
        data_rows = []
        current_year = BASE_YEAR
        previous_timestamp = None

        for parts in rows_iter:
            if not parts or all(not p.strip() for p in parts):
                continue

            if len(parts) < len(headers):
                parts = parts + [""] * (len(headers) - len(parts))
            elif len(parts) > len(headers):
                parts = parts[: len(headers)]

            timestamp, has_year = parse_timestamp_with_year_info(parts[0], current_year)
            if isinstance(timestamp, datetime):
                if has_year:
                    current_year = timestamp.year
                else:
                    while previous_timestamp is not None and timestamp < previous_timestamp:
                        current_year += 1
                        timestamp = add_years(timestamp, 1)
                previous_timestamp = timestamp

            row = [timestamp]
            for i in range(1, len(headers)):
                row.append(parse_cell(parts[i]))
            data_rows.append(row)

    return headers, data_rows

def parse_time_range(time_range: str | None, data_rows):
    if not time_range:
        return None, None

    if "~" not in time_range:
        raise ValueError(
            'Time range must use "~", for example: "Nov 02 18:57:30 ~ Jan 15 01:57:30"'
        )

    timestamps = [row[0] for row in data_rows if isinstance(row[0], datetime)]
    if not timestamps:
        raise ValueError("Cannot use -t because TIME_STAMP values were not parsed as datetime values")

    data_start = timestamps[0]
    data_end = timestamps[-1]
    start_text, end_text = (part.strip() for part in time_range.split("~", 1))
    start = parse_timestamp(start_text, BASE_YEAR) if start_text else None
    end = parse_timestamp(end_text, BASE_YEAR) if end_text else None

    if start is not None and not isinstance(start, datetime):
        raise ValueError(f"Invalid start TIME_STAMP in range: {start_text!r}")
    if end is not None and not isinstance(end, datetime):
        raise ValueError(f"Invalid end TIME_STAMP in range: {end_text!r}")

    if start is not None and end is not None:
        while end < start:
            end = add_years(end, 1)
        while end < data_start:
            start = add_years(start, 1)
            end = add_years(end, 1)
        while start > data_end:
            start = add_years(start, -1)
            end = add_years(end, -1)
    elif start is not None:
        while start < data_start:
            start = add_years(start, 1)
        while start > data_end:
            start = add_years(start, -1)
    elif end is not None:
        while end < data_start:
            end = add_years(end, 1)
        while end > data_end:
            end = add_years(end, -1)
    else:
        raise ValueError("Time range is empty")

    return start, end

def filter_data_rows_by_time(data_rows, time_range: str | None):
    start, end = parse_time_range(time_range, data_rows)
    if start is None and end is None:
        return data_rows

    filtered_rows = []
    for row in data_rows:
        timestamp = row[0]
        if not isinstance(timestamp, datetime):
            continue
        if start is not None and timestamp < start:
            continue
        if end is not None and timestamp > end:
            continue
        filtered_rows.append(row)

    if not filtered_rows:
        raise ValueError(f"No rows matched time range: {time_range}")

    return filtered_rows

def write_data_sheet(wb: Workbook, sheet_name: str, headers, data_rows):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name, 0)
    ws.append(headers)

    for row in data_rows:
        ws.append([format_timestamp_for_excel(row[0])] + row[1:])

    for cell in ws["A"][1:]:
        cell.number_format = "@"

    return ws

def find_numeric_columns(ws, header_row=1, data_start_row=2, sample_rows=50):
    numeric_cols = []
    max_scan = min(ws.max_row, data_start_row + sample_rows)

    for col in range(2, ws.max_column + 1):
        has_numeric = False
        for row in range(data_start_row, max_scan + 1):
            if is_number(ws.cell(row=row, column=col).value):
                has_numeric = True
                break
        if has_numeric:
            numeric_cols.append(col)

    return numeric_cols

def create_charts_sheet(wb, charts_sheet_name="Charts"):
    if charts_sheet_name in wb.sheetnames:
        del wb[charts_sheet_name]
    return wb.create_sheet(charts_sheet_name)

def add_chart_for_column(source_ws, target_ws, col_idx, chart_pos, header_row=1, data_start_row=2):
    title = source_ws.cell(row=header_row, column=col_idx).value or f"Column {col_idx}"
    max_row = source_ws.max_row

    chart = LineChart()
    chart.title = f"{source_ws.title} - {title}"
    chart.style = 2
    chart.height = 8
    chart.width = 14

    chart.legend = None

    chart.y_axis.title = None
    chart.y_axis.delete = False
    chart.y_axis.axId = 100

    chart.x_axis.delete = False
    chart.x_axis.title = None
    chart.x_axis.tickLblPos = "low"

    data = Reference(
        source_ws,
        min_col=col_idx,
        min_row=header_row,
        max_row=max_row,
    )
    cats = Reference(
        source_ws,
        min_col=1,
        min_row=data_start_row,
        max_row=max_row,
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    target_ws.add_chart(chart, chart_pos)

def chart_anchor_positions():
    base = [
        "A1", "J1",
        "A20", "J20",
        "A39", "J39",
        "A58", "J58",
        "A77", "J77",
    ]
    return base

def next_chart_pos(index: int, positions: list[str]) -> str:
    if index < len(positions):
        return positions[index]
    row_block = 1 + (index // 2) * 19
    col_block = "A" if index % 2 == 0 else "J"
    return f"{col_block}{row_block}"

def generate_charts_from_sheet(wb, data_sheet_name: str, max_charts: int | None = None):
    charts_ws = create_charts_sheet(wb, "Charts")
    source_ws = wb[data_sheet_name]

    if source_ws.max_row < 2 or source_ws.max_column < 2:
        return

    numeric_cols = find_numeric_columns(source_ws)
    if max_charts is not None:
        numeric_cols = numeric_cols[: max(0, max_charts)]

    positions = chart_anchor_positions()
    for pos_index, col_idx in enumerate(numeric_cols):
        chart_pos = next_chart_pos(pos_index, positions)
        add_chart_for_column(source_ws, charts_ws, col_idx, chart_pos)

def make_delta_rows(data_rows):
    delta_rows = []

    for row_index in range(1, len(data_rows)):
        current_row = data_rows[row_index]
        previous_row = data_rows[row_index - 1]
        delta_row = [format_timestamp_for_excel(current_row[0])]

        for col_index in range(1, len(current_row)):
            current_value = current_row[col_index]
            previous_value = previous_row[col_index]
            if is_number(current_value) and is_number(previous_value):
                delta_row.append(current_value - previous_value)
            else:
                delta_row.append(None)

        delta_rows.append(delta_row)

    return delta_rows

def delta_table_start_row(chart_count: int) -> int:
    if chart_count <= 0:
        return 1
    chart_row_blocks = (chart_count + 1) // 2
    return 1 + chart_row_blocks * 19 + 1

def write_delta_table(ws, headers, delta_rows, start_row: int):
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_index, value=header)

    for row_offset, row in enumerate(delta_rows, start=1):
        excel_row = start_row + row_offset
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=excel_row, column=col_index, value=value)

    for row in range(start_row + 1, start_row + len(delta_rows) + 1):
        ws.cell(row=row, column=1).number_format = "@"

def generate_delta_charts_sheet(
    wb,
    headers,
    data_rows,
    source_data_sheet_name: str,
    max_charts: int | None = None,
):
    delta_ws = create_charts_sheet(wb, "Delta Charts")

    if len(data_rows) < 2:
        delta_ws["A1"] = "Not enough rows to calculate delta values."
        return

    source_ws = wb[source_data_sheet_name]
    numeric_cols = find_numeric_columns(source_ws)
    if max_charts is not None:
        numeric_cols = numeric_cols[: max(0, max_charts)]

    table_start_row = delta_table_start_row(len(numeric_cols))
    delta_rows = make_delta_rows(data_rows)
    write_delta_table(delta_ws, headers, delta_rows, table_start_row)

    positions = chart_anchor_positions()
    for pos_index, col_idx in enumerate(numeric_cols):
        chart_pos = next_chart_pos(pos_index, positions)
        add_chart_for_column(
            delta_ws,
            delta_ws,
            col_idx,
            chart_pos,
            header_row=table_start_row,
            data_start_row=table_start_row + 1,
        )

def dump_to_excel(
    dump_path: Path,
    output_path: Path,
    sheet_name: str = "OpenSM_Statistics",
    max_charts: int | None = None,
    time_range: str | None = None,
):
    headers, data_rows = read_dump_csv(dump_path)
    if not headers:
        raise ValueError(f"No header row in {dump_path}")
    data_rows = filter_data_rows_by_time(data_rows, time_range)

    wb = Workbook()

    default = wb.active
    wb.remove(default)

    write_data_sheet(wb, sheet_name, headers, data_rows)
    generate_charts_from_sheet(wb, sheet_name, max_charts=max_charts)
    generate_delta_charts_sheet(
        wb,
        headers,
        data_rows,
        sheet_name,
        max_charts=max_charts,
    )

    wb.save(output_path)
    print(f"Saved: {output_path} ({len(data_rows)} data rows, sheet {sheet_name!r})")

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Convert opensm-statistics.dump (CSV) to Excel with line charts "
            "on a Charts sheet (openpyxl)."
        )
    )
    parser.add_argument(
        "dump_file",
        nargs="?",
        default="opensm-statistics.dump",
        help="Path to .dump CSV file (default: opensm-statistics.dump)",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output .xlsx path (default: <dump stem>_charts.xlsx next to dump)",
    )
    parser.add_argument(
        "-s",
        "--sheet",
        default="OpenSM_Statistics",
        help="Worksheet name for tabular data (default: OpenSM_Statistics)",
    )
    parser.add_argument(
        "--max-charts",
        type=int,
        default=None,
        metavar="N",
        help="Limit to first N numeric columns (after TIME_STAMP) for chart objects",
    )
    parser.add_argument(
        "-t",
        "--time-range",
        default=None,
        help='Inclusive TIME_STAMP range, for example: "Nov 02 18:57:30 ~ Jan 15 01:57:30" (default: all rows)',
    )

    args = parser.parse_args()
    dump_path = Path(args.dump_file).resolve()

    if not dump_path.exists():
        raise FileNotFoundError(f"Dump file not found: {dump_path}")

    if args.output:
        output_path = Path(args.output).resolve()
    else:
        output_path = dump_path.with_name(f"{dump_path.stem}_charts.xlsx")

    dump_to_excel(
        dump_path,
        output_path,
        sheet_name=args.sheet,
        max_charts=args.max_charts,
        time_range=args.time_range,
    )

if __name__ == "__main__":
    main()
